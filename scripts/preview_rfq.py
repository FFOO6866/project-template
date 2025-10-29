from openpyxl import load_workbook
import glob

sample_data_dir = r"C:\Users\fujif\OneDrive\Desktop\Business Development\Horme\Sample Data"
excel_files = glob.glob(f"{sample_data_dir}/*.xlsx")

if excel_files:
    sample_file = excel_files[0]
    filename = sample_file.split('\\')[-1]
    print(f"Previewing: {filename}\n")

    wb = load_workbook(sample_file)
    ws = wb.active

    print("=" * 60)
    print("RFQ DETAILS")
    print("=" * 60)
    print(f"Company: {ws['A4'].value}")
    print(f"Address: {ws['A5'].value}")
    print(f"Contact: {ws['A6'].value}")
    print(f"Phone: {ws['A7'].value}")
    print(f"Email: {ws['A8'].value}")
    print(f"\nRFQ Number: {ws['F3'].value}")
    print(f"Date: {ws['F4'].value}")
    print(f"Valid Until: {ws['F5'].value}")

    print("\n" + "=" * 60)
    print("FIRST 10 ITEMS")
    print("=" * 60)

    for row in range(18, 28):
        item_no = ws.cell(row, 1).value
        if item_no:
            product = ws.cell(row, 2).value
            category = ws.cell(row, 3).value
            qty = ws.cell(row, 4).value
            unit = ws.cell(row, 5).value
            print(f"{item_no}. {product}")
            print(f"   Category: {category}")
            print(f"   Quantity: {qty} {unit}\n")

    print("=" * 60)
    print(f"Total items in this RFQ: {len([r for r in ws.iter_rows(min_row=18, max_row=200) if r[0].value])}")
    print("=" * 60)
